import openpyxl
# Import letters for columns
from openpyxl.utils import get_column_letter
# import date cleaner and verification function
from utils import date_cleaner
# import function for capitalizing
from utils import capitalize_string, title_string
# import function to clean the microchip number
from utils import microchip_cleaner
# import function to clean weight, fetal number and nb_month
from utils import weight_cleaner
# import function to clean sex
from utils import sex_cleaner
# import age to birthdate converter and cleaner
from utils import age_birthdate_convcerter
# import translator
from utils import translate
# function separates name_surname into name and surname
from utils import name_seprator

# accessing excel
wb = openpyxl.load_workbook('excel_unprocessed/2021-5-PETVET-BAZA EXCEL.xlsx')
ws = wb.active

# prints all sheets in excel file
print('SHEET NAMES IN EXCEL:\n', wb.sheetnames)
# TODO print length of sheets
# TODO check which sheet is the correct one


# excel column names_keys
key_excel_list = ['clinic', 'nb_month', 'surgery_date', 'release_date',
                  'vet_name', 'owner_name', 'owner_address', 'dog_name',
                  'advert', 'catcher', 'feeder', 'collarID', 'sex', 'birthdate',
                  'breed', 'coat', 'weight', 'pregnancy',
                  'fetal_number', 'complications', 'rabies_vaccine', 'ear_tag',
                  'microchip', 'picture', 'comments']

# dictionary where keys will be inserted
row_dict = {}


# gets values from extracted excel row and puts them to the correspondent
def extract_row_values():
    # range (5,6): takes the row between = row 5 only!
    for row in range(5, 6):
        for i in range(25):
            row_dict[key_excel_list[i]] = ws[get_column_letter(i + 1) + str(row)].value
    return row_dict


# runs function for extraction of values from a row in excel file
extract_row_values()

# translate these keys into english from bosnian
for key in ['advert', 'pregnancy', 'complications', 'comments', 'coat']:
    if row_dict[key] != None:
        row_dict[key] = translate(row_dict[key])

# runs function for date cleaning: surgery_date
row_dict['surgery_date'] = date_cleaner(row_dict['surgery_date'])
# runs function for date cleaning: release_date
row_dict['release_date'] = date_cleaner(row_dict['release_date'])

# capitalize word in a string
for key in ['owner_name', 'dog_name', 'catcher', 'feeder', 'rabies_vaccine', 'owner_address']:
    if row_dict[key] != None:
        if key == 'owner_address':
            # capitalize only first word
            row_dict[key] = capitalize_string(row_dict[key])
        else:
            # capitalize all words
            row_dict[key] = title_string(row_dict[key])

# separates name and surname into 2 fields
#row_dict['owner_name'], row_dict['owner_surname'] = name_seprator(row_dict['owner_name'])

# adjust 15 digit microchip number and convert it to string
row_dict['microchip'] = microchip_cleaner(row_dict['microchip'])

# cleans weight to float output
row_dict['weight'] = weight_cleaner(row_dict['weight'])
# if number of fetuses is not null than converts the number to a float
if row_dict['fetal_number'] != None:
    row_dict['fetal_number'] = weight_cleaner(row_dict['fetal_number'])

# TODO number_month: so it has to start when a document is opened with the first monday in the month
#  or i onther cases where the months started.
#  It is probably to do it when we have all inserted in the SQL and than sort  it out.
# get a number in return
row_dict['nb_month'] = int(weight_cleaner(row_dict['nb_month']))

# clean sex, and return
row_dict['sex'] = sex_cleaner(row_dict['sex'])

# convert age to birthdate
# get two outputs birthdate and age
# if age not specified returns unknown for birthdate and grown for age
# the dictionary will get a new key:field row_dict['age']
row_dict['birthdate'], row_dict['age'] = age_birthdate_convcerter(row_dict['birthdate'], row_dict['surgery_date'])

# TEST PRINT
print(row_dict)
# print(type(row_dict['surgery_date']))
# print(row_dict['owner_address'])
# print(row_dict['owner_name'])
# print(row_dict['catcher'])

# TODO recognize which is the first row in excel file to be extracted (usually its the 5t row)
# print value from a specific cell test correct row access
# print('This ius the value from the cell A3 it must be "Ime klinike": {} \nThis is the value from cell A5 it must be the clinic name: {}'.format(ws['A3'].value, ws['A5'].value))

# print('WEIGHT EXTRACTED:\n', row_dict['weight'], type(row_dict['weight']))
# print('MICROCHIP EXTRACTED:\n', row_dict['microchip'], type(row_dict['microchip']), 'length:',
#       len(str(row_dict['microchip'])))
